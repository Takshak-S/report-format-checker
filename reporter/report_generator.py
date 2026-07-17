"""
reporter/report_generator.py

Generates an Excel report (.xlsx) with:
  - Summary sheet (counts by category and severity)
  - Detailed violations sheet (sortable, color-coded by severity)
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from utils.error_model import ViolationCollector
from utils.constants import Severity


# ── Color palette ─────────────────────────────────────────────────────────────
COLOR_ERROR   = "FFCCCC"  # light red
COLOR_WARNING = "FFF3CC"  # light amber
COLOR_INFO    = "CCE5FF"  # light blue
COLOR_HEADER  = "1F3864"  # dark navy
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUBHEADER = "D9E1F2"
COLOR_PASS    = "C6EFCE"  # light green

SEVERITY_COLORS = {
    Severity.CRITICAL: COLOR_ERROR,
    Severity.MAJOR:   COLOR_WARNING,
    Severity.MINOR:   COLOR_WARNING,
    Severity.WARNING: COLOR_WARNING,
    Severity.SUGGESTION: COLOR_INFO,
    Severity.INFO:    COLOR_INFO,
}

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_cell(ws, row, col, value, bold=True, bg=COLOR_HEADER, fg=COLOR_HEADER_FG, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def _data_cell(ws, row, col, value, bg=None, bold=False, wrap=False, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER
    return cell


def _auto_width(ws, min_w=10, max_w=60):
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
            max(min_w, min(max_len + 2, max_w))
        )


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: openpyxl.Workbook, collector: ViolationCollector, pdf_name: str):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Format Check Report — {pdf_name}"
    title_cell.font = Font(bold=True, size=14, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ts_cell.font = Font(italic=True, size=9, color="666666")
    ts_cell.alignment = Alignment(horizontal="center")

    # Overall summary boxes
    summary = collector.summary()
    row = 4
    for label, key, color in [
        ("Total Issues", "total", "E2EFDA"),
        ("Critical", "critical", COLOR_ERROR),
        ("Major", "major", COLOR_WARNING),
        ("Minor", "minor", COLOR_WARNING),
        ("Warnings", "warnings", COLOR_WARNING),
        ("Suggestions", "suggestions", COLOR_INFO),
        ("Info", "info", COLOR_INFO),
    ]:
        ws.merge_cells(f"A{row}:C{row}")
        _data_cell(ws, row, 1, label, bg="F2F2F2", bold=True)
        ws.merge_cells(f"D{row}:F{row}")
        _data_cell(ws, row, 4, summary[key], bg=color, bold=True, align="center")
        row += 1

    # Per-category breakdown
    row += 1
    _header_cell(ws, row, 1, "Category", bg=COLOR_HEADER)
    _header_cell(ws, row, 2, "Errors",   bg=COLOR_HEADER)
    _header_cell(ws, row, 3, "Warnings", bg=COLOR_HEADER)
    _header_cell(ws, row, 4, "Info",     bg=COLOR_HEADER)
    _header_cell(ws, row, 5, "Total",    bg=COLOR_HEADER)
    _header_cell(ws, row, 6, "Status",   bg=COLOR_HEADER)
    row += 1

    by_cat = collector.by_category()
    all_categories = sorted(by_cat.keys())

    for cat in all_categories:
        viols = by_cat[cat]
        errors   = sum(1 for v in viols if v.severity in (Severity.CRITICAL, Severity.MAJOR))
        warnings = sum(1 for v in viols if v.severity in (Severity.MINOR, Severity.WARNING))
        info     = sum(1 for v in viols if v.severity in (Severity.SUGGESTION, Severity.INFO))
        total    = len(viols)
        status   = "✗ FAIL" if errors > 0 else ("⚠ WARN" if warnings > 0 else "✓ PASS")
        status_color = COLOR_ERROR if errors else (COLOR_WARNING if warnings else COLOR_PASS)

        _data_cell(ws, row, 1, cat, bold=True)
        _data_cell(ws, row, 2, errors,   bg=COLOR_ERROR   if errors   else None, align="center")
        _data_cell(ws, row, 3, warnings, bg=COLOR_WARNING if warnings else None, align="center")
        _data_cell(ws, row, 4, info,     bg=COLOR_INFO    if info     else None, align="center")
        _data_cell(ws, row, 5, total, bold=True, align="center")
        _data_cell(ws, row, 6, status, bg=status_color, bold=True, align="center")
        row += 1

    _auto_width(ws, min_w=15)
    ws.column_dimensions["A"].width = 28
    ws.freeze_panes = "A7"


# ── Violations detail sheet ───────────────────────────────────────────────────

def _build_violations_sheet(wb: openpyxl.Workbook, collector: ViolationCollector):
    ws = wb.create_sheet("Violations")
    ws.sheet_view.showGridLines = False

    headers = ["#", "Severity", "Page", "Category", "Description", "Detail", "Location"]
    col_widths = [5, 10, 7, 20, 50, 45, 30]

    for col, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        _header_cell(ws, 1, col, hdr)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Sort: errors first, then warnings, then info; within each, by page
    severity_order = {
        Severity.CRITICAL: 0, 
        Severity.MAJOR: 1, 
        Severity.MINOR: 2,
        Severity.WARNING: 3,
        Severity.SUGGESTION: 4,
        Severity.INFO: 5
    }
    sorted_viols = sorted(
        collector.all,
        key=lambda v: (severity_order.get(v.severity, 3), v.page if v.page > 0 else 0),
    )

    for row_idx, v in enumerate(sorted_viols, start=2):
        bg = SEVERITY_COLORS.get(v.severity, "FFFFFF")
        page_str = str(v.page) if v.page > 0 else "Doc"

        _data_cell(ws, row_idx, 1, row_idx - 1, align="center")
        _data_cell(ws, row_idx, 2, v.severity, bg=bg, bold=True, align="center")
        _data_cell(ws, row_idx, 3, page_str, align="center")
        _data_cell(ws, row_idx, 4, v.category)
        _data_cell(ws, row_idx, 5, v.description, wrap=True)
        _data_cell(ws, row_idx, 6, v.detail or "", wrap=True)
        _data_cell(ws, row_idx, 7, v.location or "", wrap=True)
        ws.row_dimensions[row_idx].height = 30

    # Auto-filter on headers
    ws.auto_filter.ref = f"A1:G{len(sorted_viols) + 1}"


# ── Public API ────────────────────────────────────────────────────────────────

def generate_report(
    collector: ViolationCollector,
    pdf_path: str,
    output_path: str | None = None,
) -> Path:
    """
    Generate an Excel report for all violations in the collector.
    Returns the path to the saved .xlsx file.
    """
    pdf_name = Path(pdf_path).stem
    if output_path is None:
        output_path = str(Path(pdf_path).parent / f"{pdf_name}_format_report.xlsx")

    wb = openpyxl.Workbook()
    _build_summary_sheet(wb, collector, pdf_name)
    _build_violations_sheet(wb, collector)

    wb.save(output_path)
    return Path(output_path)
